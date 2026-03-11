import os
import requests
import pandas as pd
import json
import time
import logging
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

#creata log folder 
logs_folder = os.path.join("logs")
os.makedirs(logs_folder, exist_ok=True)

#create outputs folder 
output_folder = os.path.join("outputs", "part2")
os.makedirs(output_folder, exist_ok=True)
#######################################
# Configure logging to write to both a file AND the console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(logs_folder,'api_requests.log')),  # Saves to disk
        logging.StreamHandler(),  # Also print to console
    ],
)

logger = logging.getLogger(__name__)

######################################
class RateLimiter:
    """
    Smart rate limiter that tracks API usage.
    Uses a sliding time window to count recent requests.
    """

    def __init__(self, max_requests=60, time_window=3600):
        """
        Args:
            max_requests: Maximum requests allowed in the time window
            time_window: Time window in seconds (3600 = 1 hour)
        """
        self.max_requests = max_requests
        self.time_window = time_window
        self.requests = []  # List of timestamps of past requests

    def wait_if_needed(self):
        """Wait if we've hit the rate limit before making a new request."""
        now = time.time()

        # Remove old timestamps outside the sliding time window
        self.requests = [
            req_time for req_time in self.requests if now - req_time < self.time_window
        ]

        # If we've used up our quota, sleep until the oldest request expires
        if len(self.requests) >= self.max_requests:
            oldest_request = self.requests[0]
            sleep_time = self.time_window - (now - oldest_request)
            if sleep_time > 0:
                print(
                    f"⏰ Rate limit reached. Sleeping for {sleep_time:.1f} seconds..."
                )
                time.sleep(sleep_time)
            self.requests = []  # Clear after sleeping

        # Record the timestamp of this new request
        self.requests.append(now)

###############################################
def check_rate_limit(response):
    """
    Check rate limit info from response headers.
    GitHub includes rate limit details in every response.
    """
    if 'X-RateLimit-Limit' in response.headers:
        limit = int(response.headers['X-RateLimit-Limit'])
        remaining = int(response.headers['X-RateLimit-Remaining'])
        reset_timestamp = int(response.headers['X-RateLimit-Reset'])
        reset_time = datetime.fromtimestamp(reset_timestamp)

        print(f"Rate Limit: {remaining}/{limit}")
        print(f"Resets at: {reset_time}")

        # Warn when running low on available requests
        if remaining < 10:
            print("⚠️ Warning: Low on API requests!")

        return remaining
    return None

#######################################################
class GitHubAPI:
    """
    Reusable GitHub API client with all best practices:
    - Session management with retry logic
    - Rate limiting
    - Authentication via token
    - Logging
    """

    def __init__(self, token=None):
        self.base_url = 'https://api.github.com'
        self.session = self._create_session()  # Robust session with retries
        self.rate_limiter = RateLimiter(
            max_requests=5000, time_window=3600
        )  # Authenticated limits

        # Add authentication token if provided
        if token:
            self.session.headers.update({'Authorization': f'Bearer {token}'})

        # Always set these headers for proper API communication
        self.session.headers.update(
            {
                'Accept': 'application/vnd.github.v3+json',
                'User-Agent': 'Library-Tutorial/1.0',
            }
        )

        self.logger = logging.getLogger(self.__class__.__name__)

    def _create_session(self):
        """Create session with retry logic (private method)."""
        session = requests.Session()
        retry_strategy = Retry(
            total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session

    def get(self, endpoint, params=None):
        """Make GET request with rate limiting."""
        self.rate_limiter.wait_if_needed()  # Respect rate limits
        url = f"{self.base_url}/{endpoint.lstrip('/')}"

        try:
            response = self.session.get(url, params=params, timeout=10)
            response.raise_for_status()  # Raises exception for 4xx/5xx
            self.logger.info(f"GET {endpoint} - Status: {response.status_code}")

            # Peek at remaining rate limit with each response
            remaining = check_rate_limit(response)
            return response.json()

        except Exception as e:
            self.logger.error(f"Error fetching {endpoint}: {e}")
            raise

    def get_repo(self, owner, repo):
        """Get repository information."""
        return self.get(f'/repos/{owner}/{repo}')

    def get_user_repos(self, username):
        """Get all repositories for a user."""
        return self.get(f'/users/{username}/repos', params={'per_page': 100})

    def search_repos(self, query, language=None, min_stars=None):
        """
        Search repositories.

        Args:
            query: Search query string
            language: Filter by programming language
            min_stars: Minimum stars required

        Returns:
            list: Repository results
        """
        # Build search query by combining filters with spaces
        q_parts = [query]
        if language:
            q_parts.append(f"language:{language}")
        if min_stars:
            q_parts.append(f"stars:>={min_stars}")

        q = ' '.join(q_parts)
        results = self.get('/search/repositories', params={'q': q})
        return results['items']

    def to_dataframe(self, repos):
        """Convert repository list to DataFrame for analysis."""
        data = []
        for repo in repos:
            data.append(
                {
                    'name': repo['name'],
                    'full_name': repo['full_name'],
                    'description': repo.get('description'),
                    'stars': repo['stargazers_count'],
                    'forks': repo['forks_count'],
                    'language': repo.get('language'),
                    'created_at': repo['created_at'],
                    'updated_at': repo['updated_at'],
                }
            )
        return pd.DataFrame(data)
###############################################
#task 1.1
def task1_fetch_repos():
    """
    Fetch repository information for major ML frameworks.
    Returns a DataFrame with key metrics.
    """
    repos = ['tensorflow/tensorflow', 'pytorch/pytorch', 'scikit-learn/scikit-learn']
    repos_list=[]
    
    # Your code here
    for repo in repos:
        load_dotenv()
        api_key = os.getenv("GITHUB_TOKEN")
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Accept": "application/vnd.github.v3+json",
            "User-Agent": "Library-Tutorial-App",
        }
        url = f'https://api.github.com/repos/{repo}'
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()

            repo_info = {
                'name': data['name'],
                'stars': data['stargazers_count'],
                'forks': data['forks_count'],
                'language': data['language'],
                'open_issues': data['open_issues_count'],
                'created_at': data['created_at'],
            }
            repos_list.append(repo_info)
    
        else:
            print(f"Error: {response.status_code}")
            return None
    
    return pd.DataFrame(repos_list)

# Call the function
df = task1_fetch_repos()
df.to_csv(os.path.join(output_folder, 'task1_github.csv'), index=False)
###########################################
#task 1.2
def get_repos_metrics(df):
    df["created_at"] = pd.to_datetime(df["created_at"])
    now=pd.Timestamp.utcnow()
    df["repo_age_days"] = ( now - df["created_at"]).dt.days       
    df["stars_per_day"] = df["stars"] / df["repo_age_days"]
    df["issues_per_star_ratio"] = df["open_issues"] / df["stars"]

    df.to_csv(os.path.join(output_folder, 'task1_metrics.csv'), index=False)

get_repos_metrics(df) 
df.head()   
#########################################
#task 1.3
def plot_repo_comparison(df):

    metrics = ["stars", "forks", "open_issues"]

    df_plot = df.set_index("name")[metrics]

    df_plot.plot(kind="bar", figsize=(10,6))

    plt.title("GitHub Repository Comparison")
    plt.xlabel("Repository")
    plt.ylabel("Metric Value")
    plt.xticks(rotation=0)
    plt.legend(title="Metrics")

    plt.tight_layout()

    plt.savefig(os.path.join(output_folder,"task1_comparison.png"))
    plt.show()


plot_repo_comparison(df)
############################################
#task 2.1
def fetch_user_repos_paginated(username):
    """
    Fetch all repositories for a user with pagination.

    Args:
        username: GitHub username

    Returns:
        list: All repositories
    """
    all_repos = []
    page = 1

    # Your implementation here
    # Remember to:
    # 1. Check for empty responses  --> means you've hit the last page
    # 2. Add delays                 --> time.sleep(1) to be polite
    # 3. Handle errors              --> try/except around requests
    # 4. Log progress               --> print or logger.info(f"Page {page}")

    while True:
        logger.info(f"Page {page}")
        params = {'page': page, 'per_page': 100} 
        try:
            load_dotenv()
            api_key = os.getenv("GITHUB_TOKEN")
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Accept": "application/vnd.github.v3+json",
                "User-Agent": "Library-Tutorial-App",
            }

            url = f'https://api.github.com/users/{username}/repos'
            response = requests.get(url,headers=headers, params=params)

            if response.status_code != 200:
                logger.error(f"Error: {response.status_code}")
                break

            data = response.json()

            # If no data returned, we've gone past the last page
            if not data or len(data) == 0:
                logger.info("No more results!")
                break

            all_repos.extend(data)
            page += 1

            # Be polite - wait between requests to avoid rate limiting
            time.sleep(1)

        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return None
        
    return all_repos

all_repos = fetch_user_repos_paginated("Rawan227")
df=pd.DataFrame(all_repos)
df.to_csv(os.path.join(output_folder,"task2_all_repos.csv"), index=False)
###########################################
#task 2.2
def get_user_repos_analysis(df):
    most_used_language=df['language'].mode()[0]
    avg_stars_per_repos=df['stargazers_count'].mean()
    total_forks = df['forks_count'].sum()
    df['created_at'] = pd.to_datetime(df['created_at'])
    df['updated_at'] = pd.to_datetime(df['updated_at'])
    recent_updated=df.loc[df['updated_at'].idxmax(), 'name']
    oldest_repo=df.loc[df['created_at'].idxmax(),'name']


    report = f"""
    Repository Analysis Report

    Most used programming language: {most_used_language}
    Average stars per repository: {avg_stars_per_repos:.2f}
    Total forks across all repos: {total_forks}
    Most recently updated repo: {recent_updated}
    Oldest repo:{oldest_repo}
    """
    with open(os.path.join(output_folder,"task2_analysis.txt"), "w") as file:
        file.write(report)

get_user_repos_analysis(df)

##################################################
#task 3
class GitHubAnalyzer(GitHubAPI):
    """
    Complete GitHub API client with analysis capabilities.
    Build on top of the GitHubAPI class concepts from section 2.7.
    """

    def __init__(self, token=None):
        # Your initialization
        # Hint: Set up session, rate_limiter, logger like in GitHubAPI
        super().__init__(token=token)
        self.logger.info("GitHubAnalyzer initialized")



    def search_repos(self, query, language=None, min_stars=0):
        """
        Search repositories with filters.

        Returns:
            DataFrame with results
        """
        repos = super().search_repos(query, language, min_stars)
        df = self.to_dataframe(repos)
        self.logger.info(f"Found {len(df)} repositories")
        return df

    def compare_repos(self, repo_list):
        """
        Compare multiple repositories.

        Args:
            repo_list: List of "owner/repo" strings

        Returns:
            DataFrame with comparison
        """
        data=[]
        for repo in repo_list:

            owner, repo_name = repo.split("/")

            repo_data = self.get_repo(owner, repo_name)

            data.append(
                {
                    "name": repo_data["name"],
                    "owner": repo_data["owner"]["login"],
                    "stars": repo_data["stargazers_count"],
                    "forks": repo_data["forks_count"],
                    "watchers": repo_data["watchers_count"],
                    "open_issues": repo_data["open_issues_count"],
                    "language": repo_data["language"],
                    "created_at": repo_data["created_at"],
                }
            )

        df = pd.DataFrame(data)
        return df 

    def export_to_excel(self, df, filename):
        """
        Export DataFrame to Excel with formatting.
        - Bold headers
        - Auto-adjust column widths
        - Add creation timestamp
        """

        df["generated_at"] = datetime.now()

        df.to_excel(filename, index=False)

        wb = load_workbook(filename)
        ws = wb.active

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)

        self.logger.info(f"Exported results to {filename}")


# Test your class by:
analyzer = GitHubAnalyzer(token=os.getenv("GITHUB_TOKEN"))

# 1. Searching for "data science" repos in Python with >500 stars
df_search = analyzer.search_repos(
    query="data science",
    language="python",
    min_stars=500
)
df_search.head()

# 2. Comparing 5 repos of your choice
repos = [
    "pandas-dev/pandas",
    "numpy/numpy",
    "scikit-learn/scikit-learn",
    "pytorch/pytorch",
    "tensorflow/tensorflow"
]
df_compare = analyzer.compare_repos(repos)
df_compare

# 3. Exporting results to task3_results.xlsx
analyzer.export_to_excel(df_compare, os.path.join(output_folder,"task3_results.xlsx"))